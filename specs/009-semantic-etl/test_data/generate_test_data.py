#!/usr/bin/env python3
"""
Test Data Generator for Semantic ETL Pipeline
==============================================

Generates Excel test files with known products for E2E testing.

Phase 9: Semantic ETL Pipeline Refactoring
Tasks: T048, T052
"""

import random
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Define realistic product categories with variations for fuzzy matching tests
CATEGORIES = {
    "Electronics": {
        "subcategories": ["Laptops", "Smartphones", "Tablets", "Accessories"],
        "brands": ["Samsung", "Apple", "Xiaomi", "Huawei", "Lenovo", "ASUS"],
        "product_templates": [
            "{brand} {model} {size}\" {spec}",
            "{brand} {model} ({spec})",
            "{brand} {model} {color}",
        ],
    },
    "Home Appliances": {
        "subcategories": ["Kitchen", "Cleaning", "Climate Control", "Laundry"],
        "brands": ["Bosch", "Samsung", "LG", "Philips", "Dyson", "Xiaomi"],
        "product_templates": [
            "{brand} {model} {type}",
            "{brand} {type} {model}",
            "{brand} {type} ({spec})",
        ],
    },
    "Tools": {
        "subcategories": ["Power Tools", "Hand Tools", "Garden", "Safety"],
        "brands": ["Makita", "Bosch", "DeWalt", "Milwaukee", "Metabo"],
        "product_templates": [
            "{brand} {type} {model}",
            "{brand} {model} {type}",
            "{brand} Professional {type}",
        ],
    },
    "Office": {
        "subcategories": ["Furniture", "Supplies", "Printers", "Accessories"],
        "brands": ["Canon", "HP", "Epson", "Logitech", "Dell"],
        "product_templates": [
            "{brand} {model} {type}",
            "{brand} {type} {model}",
            "{type} {brand} {model}",
        ],
    },
}

# Category name variations for fuzzy matching tests
CATEGORY_VARIATIONS = {
    "Electronics": ["Электроника", "Electronic", "Electronics/Gadgets"],
    "Laptops": ["Notebook", "Ноутбуки", "Laptop Computer"],
    "Smartphones": ["Смартфоны", "Mobile Phones", "Phones"],
    "Home Appliances": ["Бытовая техника", "Appliances", "Home & Kitchen"],
    "Tools": ["Инструменты", "Power Tools", "Hardware"],
}

# Product specs for variety
SIZES = ["13", "14", "15.6", "17", "27", "32", "55", "65"]
COLORS = ["Black", "White", "Silver", "Gray", "Blue", "Red"]
SPECS = ["Pro", "Max", "Ultra", "Plus", "Basic", "Premium", "Lite", "Air"]
MODELS = [f"X{i}" for i in range(1, 50)] + [f"Pro{i}" for i in range(1, 20)] + [f"S{i}" for i in range(1, 30)]


def generate_product_name(category: str, subcat: str) -> tuple[str, str]:
    """Generate a realistic product name with category path."""
    cat_info = CATEGORIES[category]
    brand = random.choice(cat_info["brands"])
    template = random.choice(cat_info["product_templates"])
    model = random.choice(MODELS)
    
    name = template.format(
        brand=brand,
        model=model,
        type=subcat,
        size=random.choice(SIZES),
        color=random.choice(COLORS),
        spec=random.choice(SPECS),
    )
    
    # Sometimes use category variations for fuzzy matching tests
    final_category = category
    if random.random() < 0.3:  # 30% use variations
        variations = CATEGORY_VARIATIONS.get(category, [])
        if variations:
            final_category = random.choice(variations)
    
    final_subcat = subcat
    if random.random() < 0.3:
        variations = CATEGORY_VARIATIONS.get(subcat, [])
        if variations:
            final_subcat = random.choice(variations)
    
    category_path = f"{final_category} > {final_subcat}"
    return name, category_path


def generate_price() -> tuple[Decimal, Decimal]:
    """Generate realistic RRC and OPT prices."""
    # Base price between 10 and 5000 BYN
    base = random.uniform(10, 5000)
    price_rrc = round(Decimal(str(base)), 2)
    
    # OPT price is typically 15-30% lower
    discount = random.uniform(0.15, 0.30)
    price_opt = round(price_rrc * Decimal(str(1 - discount)), 2)
    
    return price_rrc, price_opt


def generate_description(name: str, category: str) -> str:
    """Generate a product description."""
    adjectives = ["High quality", "Professional", "Premium", "Reliable", "Durable"]
    features = [
        "with warranty",
        "fast shipping",
        "official distributor",
        "best price",
        "new model",
    ]
    
    return f"{random.choice(adjectives)} {name}. {random.choice(features).capitalize()}."


def create_standard_test_file(output_path: Path, num_products: int = 300, duplicate_count: int = 10) -> dict:
    """
    Create standard test Excel file with known products.
    
    Args:
        output_path: Path to save Excel file
        num_products: Number of unique products to generate
        duplicate_count: Number of intentional duplicates to add
    
    Returns:
        Dictionary with test metadata for validation
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload to site"  # Priority sheet name
    
    # Define headers
    headers = ["Name", "Description", "Price RRC", "Price OPT", "Category"]
    
    # Style headers
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Track metadata for validation
    metadata = {
        "total_products": num_products,
        "categories": set(),
        "duplicate_count": 0,
        "duplicates": [],  # Track actual duplicates for test validation
        "products": [],
    }
    
    # Generate unique products first
    generated_names = set()
    row = 2
    unique_products = []
    
    while len(unique_products) < num_products:
        category = random.choice(list(CATEGORIES.keys()))
        subcat = random.choice(CATEGORIES[category]["subcategories"])
        
        name, category_path = generate_product_name(category, subcat)
        price_rrc, price_opt = generate_price()
        description = generate_description(name, category)
        
        # Ensure unique products
        if name in generated_names:
            continue
        
        generated_names.add(name)
        
        product = {
            "row": row,
            "name": name,
            "description": description,
            "price_rrc": float(price_rrc),
            "price_opt": float(price_opt),
            "category_path": category_path,
        }
        unique_products.append(product)
        
        # Write to Excel
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=description)
        ws.cell(row=row, column=3, value=float(price_rrc))
        ws.cell(row=row, column=4, value=float(price_opt))
        ws.cell(row=row, column=5, value=category_path)
        
        # Extract categories for tracking
        parts = category_path.split(" > ")
        for part in parts:
            metadata["categories"].add(part.strip())
        
        row += 1
    
    metadata["products"] = unique_products
    
    # Add intentional duplicates at the end (for deduplication testing)
    # Pick random products to duplicate with prices within 1% tolerance
    products_to_duplicate = random.sample(unique_products, min(duplicate_count, len(unique_products)))
    
    for dup_source in products_to_duplicate:
        # Same product with price within 1% tolerance (should be detected as duplicate)
        price_variation = Decimal(str(random.uniform(0.995, 1.005)))
        dup_price = round(Decimal(str(dup_source["price_rrc"])) * price_variation, 2)
        
        ws.cell(row=row, column=1, value=dup_source["name"])
        ws.cell(row=row, column=2, value=dup_source["description"])
        ws.cell(row=row, column=3, value=float(dup_price))
        ws.cell(row=row, column=4, value=float(dup_source["price_opt"]))
        ws.cell(row=row, column=5, value=dup_source["category_path"])
        
        metadata["duplicates"].append({
            "row": row,
            "name": dup_source["name"],
            "original_row": dup_source["row"],
            "original_price": dup_source["price_rrc"],
            "duplicate_price": float(dup_price),
        })
        
        metadata["duplicate_count"] += 1
        row += 1
    
    # Adjust column widths
    column_widths = [50, 60, 12, 12, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Convert categories set to list for JSON serialization
    metadata["categories"] = list(metadata["categories"])
    metadata["total_rows"] = row - 2  # Excluding header (includes duplicates)
    metadata["unique_count"] = len(unique_products)
    
    wb.save(output_path)
    print(f"Created test file: {output_path}")
    print(f"  - Unique products: {len(unique_products)}")
    print(f"  - Intentional duplicates: {metadata['duplicate_count']}")
    print(f"  - Total rows: {metadata['total_rows']}")
    print(f"  - Categories: {len(metadata['categories'])}")
    
    return metadata


def create_performance_test_file(output_path: Path, num_products: int = 500, duplicate_count: int = 15) -> dict:
    """
    Create larger test file for performance testing.
    
    Args:
        output_path: Path to save Excel file
        num_products: Number of products (default 500 for <3 min test)
        duplicate_count: Number of intentional duplicates
    
    Returns:
        Dictionary with test metadata
    """
    return create_standard_test_file(output_path, num_products, duplicate_count)


def create_multi_sheet_test_file(output_path: Path) -> dict:
    """
    Create multi-sheet Excel file for US2 testing (T061).
    
    Contains 5 sheets:
    - "Instructions" (metadata, should be skipped)
    - "Products" (product data sheet)
    - "Pricing" (product data sheet with different products)
    - "Config" (metadata, should be skipped)
    - "Upload to site" (priority sheet - should be the ONLY one processed)
    
    Args:
        output_path: Path to save Excel file
    
    Returns:
        Dictionary with test metadata for validation
    """
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metadata_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    metadata = {
        "sheets": {},
        "expected_selected": ["Upload to site"],
        "expected_skipped": ["Instructions", "Products", "Pricing", "Config"],
        "priority_sheet": "Upload to site",
        "cross_sheet_duplicates": [],
    }
    
    # Sheet 1: Instructions (metadata - should be skipped)
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    instructions_content = [
        ["Supplier File Instructions"],
        [""],
        ["1. Fill in the 'Upload to site' sheet with products"],
        ["2. Include: Name, Description, Price, Category"],
        ["3. Do not modify this sheet"],
        [""],
        ["Contact: support@marketbel.by"],
    ]
    for row_idx, row_data in enumerate(instructions_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
    
    metadata["sheets"]["Instructions"] = {
        "type": "metadata",
        "row_count": len(instructions_content),
        "should_skip": True,
    }
    
    # Sheet 2: Products (product data - but should be skipped due to priority sheet)
    ws_products = wb.create_sheet("Products")
    products_headers = ["Name", "Description", "Price RRC", "Price OPT", "Category"]
    for col, header in enumerate(products_headers, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    products_data = []
    generated_names = set()
    for row in range(2, 102):  # 100 products
        category = random.choice(list(CATEGORIES.keys()))
        subcat = random.choice(CATEGORIES[category]["subcategories"])
        name, category_path = generate_product_name(category, subcat)
        
        # Ensure unique within sheet
        attempt = 0
        while name in generated_names and attempt < 10:
            name, category_path = generate_product_name(category, subcat)
            attempt += 1
        generated_names.add(name)
        
        price_rrc, price_opt = generate_price()
        description = generate_description(name, category)
        
        ws_products.cell(row=row, column=1, value=name)
        ws_products.cell(row=row, column=2, value=description)
        ws_products.cell(row=row, column=3, value=float(price_rrc))
        ws_products.cell(row=row, column=4, value=float(price_opt))
        ws_products.cell(row=row, column=5, value=category_path)
        
        products_data.append({
            "name": name,
            "price_rrc": float(price_rrc),
        })
    
    metadata["sheets"]["Products"] = {
        "type": "product_data",
        "row_count": 101,  # Header + 100 data rows
        "product_count": 100,
        "should_skip": True,  # Skipped due to priority sheet
    }
    
    # Sheet 3: Pricing (product data - but should be skipped due to priority sheet)
    ws_pricing = wb.create_sheet("Pricing")
    pricing_headers = ["Item Name", "RRC", "Wholesale", "Category Path"]
    for col, header in enumerate(pricing_headers, 1):
        cell = ws_pricing.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    pricing_data = []
    for row in range(2, 52):  # 50 products
        category = random.choice(list(CATEGORIES.keys()))
        subcat = random.choice(CATEGORIES[category]["subcategories"])
        name, category_path = generate_product_name(category, subcat)
        
        # Ensure unique within sheet
        attempt = 0
        while name in generated_names and attempt < 10:
            name, category_path = generate_product_name(category, subcat)
            attempt += 1
        generated_names.add(name)
        
        price_rrc, price_opt = generate_price()
        
        ws_pricing.cell(row=row, column=1, value=name)
        ws_pricing.cell(row=row, column=2, value=float(price_rrc))
        ws_pricing.cell(row=row, column=3, value=float(price_opt))
        ws_pricing.cell(row=row, column=4, value=category_path)
        
        pricing_data.append({
            "name": name,
            "price_rrc": float(price_rrc),
        })
    
    # Add some cross-sheet duplicates (same products in Pricing as Products)
    duplicate_count = 5
    for i, dup_source in enumerate(random.sample(products_data, duplicate_count)):
        row = 52 + i
        ws_pricing.cell(row=row, column=1, value=dup_source["name"])
        ws_pricing.cell(row=row, column=2, value=dup_source["price_rrc"])
        ws_pricing.cell(row=row, column=3, value=dup_source["price_rrc"] * 0.8)
        ws_pricing.cell(row=row, column=4, value="Electronics > Misc")
        
        metadata["cross_sheet_duplicates"].append({
            "name": dup_source["name"],
            "sheets": ["Products", "Pricing"],
        })
    
    metadata["sheets"]["Pricing"] = {
        "type": "product_data",
        "row_count": 56,  # Header + 50 + 5 duplicates
        "product_count": 55,
        "should_skip": True,  # Skipped due to priority sheet
    }
    
    # Sheet 4: Config (metadata - should be skipped)
    ws_config = wb.create_sheet("Config")
    config_content = [
        ["Configuration Settings"],
        [""],
        ["Setting", "Value"],
        ["Supplier ID", "12345"],
        ["Currency", "BYN"],
        ["Tax Rate", "20%"],
        ["Active", "Yes"],
    ]
    for row_idx, row_data in enumerate(config_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
    
    metadata["sheets"]["Config"] = {
        "type": "metadata",
        "row_count": len(config_content),
        "should_skip": True,
    }
    
    # Sheet 5: Upload to site (PRIORITY SHEET - should be the ONLY one processed)
    ws_upload = wb.create_sheet("Upload to site")
    upload_headers = ["Название", "Описание", "Цена РРЦ", "Цена ОПТ", "Категория"]
    for col, header in enumerate(upload_headers, 1):
        cell = ws_upload.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    upload_data = []
    for row in range(2, 82):  # 80 products
        category = random.choice(list(CATEGORIES.keys()))
        subcat = random.choice(CATEGORIES[category]["subcategories"])
        name, category_path = generate_product_name(category, subcat)
        
        # Ensure unique within sheet
        attempt = 0
        while name in generated_names and attempt < 10:
            name, category_path = generate_product_name(category, subcat)
            attempt += 1
        generated_names.add(name)
        
        price_rrc, price_opt = generate_price()
        description = generate_description(name, category)
        
        ws_upload.cell(row=row, column=1, value=name)
        ws_upload.cell(row=row, column=2, value=description)
        ws_upload.cell(row=row, column=3, value=float(price_rrc))
        ws_upload.cell(row=row, column=4, value=float(price_opt))
        ws_upload.cell(row=row, column=5, value=category_path)
        
        upload_data.append({
            "name": name,
            "price_rrc": float(price_rrc),
        })
    
    metadata["sheets"]["Upload to site"] = {
        "type": "priority_sheet",
        "row_count": 81,  # Header + 80 data rows
        "product_count": 80,
        "should_process": True,
    }
    
    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 30
    
    metadata["total_sheets"] = 5
    metadata["expected_product_count"] = 80  # Only from "Upload to site"
    
    wb.save(output_path)
    print(f"Created multi-sheet test file: {output_path}")
    print(f"  - Total sheets: {metadata['total_sheets']}")
    print(f"  - Expected selected: {metadata['expected_selected']}")
    print(f"  - Expected skipped: {metadata['expected_skipped']}")
    print(f"  - Expected products (from priority sheet only): {metadata['expected_product_count']}")
    
    return metadata


if __name__ == "__main__":
    # Create test data directory if needed
    test_data_dir = Path(__file__).parent
    
    # Generate standard 300-row test file (T048)
    standard_file = test_data_dir / "standard_supplier_300rows.xlsx"
    standard_metadata = create_standard_test_file(standard_file, 300)
    
    # Generate performance test file (T052)
    perf_file = test_data_dir / "performance_test_500rows.xlsx"
    perf_metadata = create_performance_test_file(perf_file, 500)
    
    # Generate multi-sheet test file (T061 - US2)
    multi_sheet_file = test_data_dir / "multi_sheet_supplier.xlsx"
    multi_sheet_metadata = create_multi_sheet_test_file(multi_sheet_file)
    
    # Save metadata for validation
    import json
    
    metadata_file = test_data_dir / "test_metadata.json"
    with open(metadata_file, "w") as f:
        json.dump({
            "standard_file": {
                "path": str(standard_file),
                "total_products": standard_metadata["total_rows"],
                "categories": standard_metadata["categories"],
                "duplicate_count": standard_metadata["duplicate_count"],
            },
            "performance_file": {
                "path": str(perf_file),
                "total_products": perf_metadata["total_rows"],
                "categories": perf_metadata["categories"],
                "duplicate_count": perf_metadata["duplicate_count"],
            },
            "multi_sheet_file": {
                "path": str(multi_sheet_file),
                "total_sheets": multi_sheet_metadata["total_sheets"],
                "expected_selected": multi_sheet_metadata["expected_selected"],
                "expected_skipped": multi_sheet_metadata["expected_skipped"],
                "expected_product_count": multi_sheet_metadata["expected_product_count"],
                "priority_sheet": multi_sheet_metadata["priority_sheet"],
            },
        }, f, indent=2)
    
    print(f"\nMetadata saved to: {metadata_file}")

