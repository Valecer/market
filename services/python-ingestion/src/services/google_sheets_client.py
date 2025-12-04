"""Google Sheets client for exporting spreadsheets.

Phase 9: Courier pattern - only exports Google Sheets to XLSX format.
All parsing/extraction is handled by ml-analyze service.

This is NOT a parser - it's a download/export client for the courier pattern.
"""

import io
from typing import Optional
from urllib.parse import urlparse

import gspread
from gspread.exceptions import SpreadsheetNotFound, WorksheetNotFound, APIError
import structlog

try:
    from openpyxl import Workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.config import settings
from src.errors.exceptions import ParserError

logger = structlog.get_logger(__name__)


class GoogleSheetsClient:
    """
    Google Sheets client for the courier pattern.

    Only exports spreadsheets to XLSX format for ml-analyze processing.
    No parsing logic - that's handled by ml-analyze's SmartParserService.
    """

    def __init__(self, credentials_path: Optional[str] = None):
        """
        Initialize Google Sheets client with authentication.

        Args:
            credentials_path: Path to service account credentials JSON file
        """
        self.credentials_path = credentials_path or settings.google_credentials_path
        self._client: Optional[gspread.Client] = None

        try:
            self._client = gspread.service_account(filename=self.credentials_path)
            logger.info(
                "google_sheets_client_initialized",
                credentials_path=self.credentials_path,
            )
        except FileNotFoundError as e:
            raise ParserError(
                f"Google credentials file not found: {self.credentials_path}"
            ) from e
        except Exception as e:
            raise ParserError(
                f"Failed to authenticate with Google Sheets API: {e}"
            ) from e

    def _open_spreadsheet_by_url(self, sheet_url: str) -> gspread.Spreadsheet:
        """Open spreadsheet by URL."""
        log = logger.bind(sheet_url=sheet_url)

        try:
            parsed_url = urlparse(sheet_url)
            path_parts = parsed_url.path.split("/")

            if "d" in path_parts:
                spreadsheet_id = path_parts[path_parts.index("d") + 1]
            else:
                raise ParserError(f"Invalid Google Sheets URL format: {sheet_url}")

            spreadsheet = self._client.open_by_key(spreadsheet_id)
            log.debug(
                "spreadsheet_opened",
                spreadsheet_id=spreadsheet_id,
                title=spreadsheet.title,
            )
            return spreadsheet

        except Exception as e:
            raise ParserError(f"Failed to open spreadsheet: {e}") from e

    def _get_worksheet(
        self, spreadsheet: gspread.Spreadsheet, sheet_name: Optional[str]
    ) -> gspread.Worksheet:
        """Get worksheet by name with fallback to first sheet."""
        log = logger.bind(sheet_name=sheet_name)

        if not sheet_name:
            worksheets = spreadsheet.worksheets()
            if not worksheets:
                raise ParserError("No worksheets found in spreadsheet")
            return worksheets[0]

        try:
            worksheet = spreadsheet.worksheet(sheet_name)
            log.debug("worksheet_found", sheet_name=sheet_name)
            return worksheet
        except WorksheetNotFound:
            available = [ws.title for ws in spreadsheet.worksheets()]
            if not available:
                raise ParserError("No worksheets found in spreadsheet")

            first_sheet = available[0]
            log.warning(
                "worksheet_not_found_using_fallback",
                requested_sheet=sheet_name,
                fallback_sheet=first_sheet,
                available_sheets=available,
            )
            return spreadsheet.worksheet(first_sheet)

    async def export_to_xlsx(
        self,
        sheet_url: str,
        sheet_name: Optional[str] = None,
    ) -> bytes:
        """
        Export Google Sheet to XLSX format.

        Downloads the entire spreadsheet (or specified sheet) and converts
        it to XLSX format for ML processing by ml-analyze service.

        Args:
            sheet_url: Google Sheets URL
            sheet_name: Optional specific sheet to export (defaults to all)

        Returns:
            XLSX file as bytes

        Raises:
            ParserError: If export fails
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Install with: pip install openpyxl"
            )

        log = logger.bind(sheet_url=sheet_url, sheet_name=sheet_name)
        log.info("export_to_xlsx_started")

        try:
            spreadsheet = self._open_spreadsheet_by_url(sheet_url)

            # Create workbook
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            if default_sheet is not None:
                wb.remove(default_sheet)

            # Determine which sheets to export
            if sheet_name:
                worksheets = [self._get_worksheet(spreadsheet, sheet_name)]
            else:
                worksheets = spreadsheet.worksheets()

            for ws in worksheets:
                # Create sheet in workbook (Excel sheet name limit is 31 chars)
                xlsx_ws = wb.create_sheet(title=ws.title[:31])

                # Get all values
                all_values = ws.get_all_values()

                # Write values to xlsx
                for row_idx, row_data in enumerate(all_values, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        xlsx_ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            xlsx_bytes = output.getvalue()

            log.info(
                "export_to_xlsx_completed",
                sheets_exported=len(worksheets),
                size_bytes=len(xlsx_bytes),
            )

            return xlsx_bytes

        except (SpreadsheetNotFound, WorksheetNotFound) as e:
            raise ParserError(f"Sheet not found: {e}") from e
        except APIError as e:
            raise ParserError(f"Google Sheets API error during export: {e}") from e
        except Exception as e:
            raise ParserError(f"Failed to export to XLSX: {e}") from e

