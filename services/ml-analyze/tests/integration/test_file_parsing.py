"""
Integration Tests for File Parsing
===================================

End-to-end tests for the file parsing pipeline.
Tests actual file parsing without database operations.
"""

from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pytest

from src.ingest.parser_factory import ParserFactory, get_parser, get_parser_for_file
from src.schemas.domain import NormalizedRow
from src.services.ingestion_service import IngestionService


@pytest.fixture
def sample_excel_path(tmp_path: Path) -> Path:
    """Create a sample Excel file for testing."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Header row
        ws["A1"] = "Product Name"
        ws["B1"] = "Price"
        ws["C1"] = "SKU"
        ws["D1"] = "Category"
        ws["E1"] = "Brand"

        # Data rows
        ws["A2"] = "Energizer AA Batteries 24-Pack"
        ws["B2"] = 19.99
        ws["C2"] = "EN-AA-24"
        ws["D2"] = "Batteries"
        ws["E2"] = "Energizer"

        ws["A3"] = "Duracell AA Batteries 12-Pack"
        ws["B3"] = 12.99
        ws["C3"] = "DU-AA-12"
        ws["D3"] = "Batteries"
        ws["E3"] = "Duracell"

        ws["A4"] = "USB-C Cable 6ft"
        ws["B4"] = 9.99
        ws["C4"] = "USB-C-6"
        ws["D4"] = "Cables"
        ws["E4"] = "Generic"

        file_path = tmp_path / "products.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path
    except ImportError:
        pytest.skip("openpyxl not installed")


@pytest.fixture
def merged_cell_excel(tmp_path: Path) -> Path:
    """Create Excel with merged cells."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Headers
        ws["A1"] = "Category"
        ws["B1"] = "Product"
        ws["C1"] = "Price"

        # Category A merged
        ws["A2"] = "Electronics"
        ws.merge_cells("A2:A4")

        ws["B2"] = "Smartphone"
        ws["C2"] = 999.99

        ws["B3"] = "Tablet"
        ws["C3"] = 599.99

        ws["B4"] = "Laptop"
        ws["C4"] = 1299.99

        # Category B merged
        ws["A5"] = "Office"
        ws.merge_cells("A5:A6")

        ws["B5"] = "Desk"
        ws["C5"] = 299.99

        ws["B6"] = "Chair"
        ws["C6"] = 199.99

        file_path = tmp_path / "merged.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path
    except ImportError:
        pytest.skip("openpyxl not installed")


class TestParserFactoryIntegration:
    """Integration tests for ParserFactory."""

    def test_create_excel_parser(self) -> None:
        """Should create Excel parser."""
        parser = ParserFactory.create("excel")
        assert parser.can_handle("file.xlsx")

    def test_create_pdf_parser(self) -> None:
        """Should create PDF parser."""
        parser = ParserFactory.create("pdf")
        assert parser.can_handle("file.pdf")

    def test_from_file_path_xlsx(self, sample_excel_path: Path) -> None:
        """Should detect Excel parser from .xlsx path."""
        parser = ParserFactory.from_file_path(sample_excel_path)
        assert parser.can_handle(sample_excel_path)

    def test_get_supported_types(self) -> None:
        """Should return all supported types."""
        types = ParserFactory.get_supported_types()
        assert "pdf" in types
        assert "xlsx" in types
        assert "excel" in types


class TestExcelParsingIntegration:
    """Integration tests for Excel parsing."""

    @pytest.mark.asyncio
    async def test_parse_simple_excel(self, sample_excel_path: Path) -> None:
        """Should parse simple Excel file."""
        parser = get_parser("excel")
        rows = await parser.parse(sample_excel_path)

        assert len(rows) == 3
        assert all(isinstance(r, NormalizedRow) for r in rows)

    @pytest.mark.asyncio
    async def test_parse_extracts_all_fields(self, sample_excel_path: Path) -> None:
        """Should extract all mapped fields."""
        parser = get_parser_for_file(sample_excel_path)
        rows = await parser.parse(sample_excel_path)

        first = rows[0]
        assert first.name == "Energizer AA Batteries 24-Pack"
        assert first.price == Decimal("19.99")
        assert first.sku == "EN-AA-24"
        assert first.category == "Batteries"
        assert first.brand == "Energizer"

    @pytest.mark.asyncio
    async def test_parse_merged_cells_forward_fill(
        self,
        merged_cell_excel: Path,
    ) -> None:
        """Should forward-fill merged cells."""
        parser = get_parser("excel")
        rows = await parser.parse(merged_cell_excel)

        # Should have 5 rows total
        assert len(rows) == 5

        # First 3 rows should have "Electronics" category
        electronics_rows = [r for r in rows if r.category == "Electronics"]
        assert len(electronics_rows) == 3

        # Last 2 rows should have "Office" category
        office_rows = [r for r in rows if r.category == "Office"]
        assert len(office_rows) == 2


class TestPdfParsingIntegration:
    """Integration tests for PDF parsing."""

    @pytest.mark.asyncio
    async def test_parse_pdf_with_mock(self, tmp_path: Path) -> None:
        """Should parse PDF using mocked pymupdf4llm."""
        pdf_file = tmp_path / "test.pdf"
        pdf_file.write_bytes(b"%PDF-1.4 fake content")

        markdown = """
| Product | Price | SKU |
|---------|-------|-----|
| Item 1  | 10.00 | A01 |
| Item 2  | 20.00 | A02 |
"""

        with patch("src.ingest.pdf_strategy.pymupdf4llm") as mock_pdf:
            mock_pdf.to_markdown.return_value = markdown

            parser = get_parser("pdf")
            rows = await parser.parse(pdf_file)

            assert len(rows) == 2
            assert rows[0].price == Decimal("10.00")
            assert rows[1].price == Decimal("20.00")


class TestIngestionServiceIntegration:
    """Integration tests for IngestionService."""

    @pytest.mark.asyncio
    async def test_parse_only(self, sample_excel_path: Path) -> None:
        """Should parse file without database insertion."""
        service = IngestionService()
        rows = await service.parse_only(sample_excel_path)

        assert len(rows) == 3
        assert all(isinstance(r, NormalizedRow) for r in rows)

    @pytest.mark.asyncio
    async def test_validate_file(self, sample_excel_path: Path) -> None:
        """Should validate parseable files."""
        service = IngestionService()

        is_valid = await service.validate_file(sample_excel_path)
        assert is_valid is True

    @pytest.mark.asyncio
    async def test_validate_nonexistent(self) -> None:
        """Should reject nonexistent files."""
        service = IngestionService()

        is_valid = await service.validate_file("/nonexistent/file.xlsx")
        assert is_valid is False


class TestChunkingIntegration:
    """Integration tests for chunking parsed rows."""

    @pytest.mark.asyncio
    async def test_chunk_parsed_rows(self, sample_excel_path: Path) -> None:
        """Should chunk parsed rows."""
        from src.ingest.chunker import Chunker

        parser = get_parser("excel")
        rows = await parser.parse(sample_excel_path)

        chunker = Chunker()
        chunks = chunker.chunk_batch(rows)

        assert len(chunks) == 3
        assert all(c.text for c in chunks)

        # Check first chunk contains product info
        first_chunk = chunks[0]
        assert "Energizer" in first_chunk.text
        assert "AA" in first_chunk.text or "Batteries" in first_chunk.text

    @pytest.mark.asyncio
    async def test_chunk_metadata(self, sample_excel_path: Path) -> None:
        """Chunks should have metadata from source rows."""
        from src.ingest.chunker import Chunker

        parser = get_parser("excel")
        rows = await parser.parse(sample_excel_path)

        chunker = Chunker()
        chunks = chunker.chunk_batch(rows)

        # Check metadata
        first_chunk = chunks[0]
        assert first_chunk.metadata.get("category") == "Batteries"
        assert first_chunk.metadata.get("brand") == "Energizer"
        assert first_chunk.metadata.get("price") == 19.99

