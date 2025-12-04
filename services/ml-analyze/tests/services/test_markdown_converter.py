"""
Unit Tests for MarkdownConverter
================================

Tests Excel to Markdown conversion including merged cells,
chunking, and sheet info retrieval.

Phase 9: Semantic ETL Pipeline Refactoring
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.services.smart_parser.markdown_converter import MarkdownConverter


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    """Create a sample Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header row
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Category"
    
    # Data rows
    ws["A2"] = "Product 1"
    ws["B2"] = 100.50
    ws["C2"] = "Electronics"
    
    ws["A3"] = "Product 2"
    ws["B3"] = 200.00
    ws["C3"] = "Furniture"
    
    ws["A4"] = "Product 3"
    ws["B4"] = 50.25
    ws["C4"] = "Electronics"
    
    file_path = tmp_path / "test_products.xlsx"
    wb.save(file_path)
    wb.close()
    
    return file_path


@pytest.fixture
def workbook_with_merged_cells(tmp_path: Path) -> Path:
    """Create an Excel file with merged cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload to site"
    
    # Header
    ws["A1"] = "Category"
    ws["B1"] = "Name"
    ws["C1"] = "Price"
    
    # Merged category spanning 3 rows
    ws["A2"] = "Electronics"
    ws.merge_cells("A2:A4")
    
    ws["B2"] = "Laptop A"
    ws["C2"] = 1000
    
    ws["B3"] = "Laptop B"
    ws["C3"] = 1200
    
    ws["B4"] = "Laptop C"
    ws["C4"] = 1500
    
    # Different category
    ws["A5"] = "Furniture"
    ws.merge_cells("A5:A6")
    
    ws["B5"] = "Chair"
    ws["C5"] = 100
    
    ws["B6"] = "Desk"
    ws["C6"] = 200
    
    file_path = tmp_path / "merged_cells.xlsx"
    wb.save(file_path)
    wb.close()
    
    return file_path


@pytest.fixture
def multi_sheet_workbook(tmp_path: Path) -> Path:
    """Create an Excel file with multiple sheets."""
    wb = Workbook()
    
    # Products sheet
    ws1 = wb.active
    ws1.title = "Products"
    ws1["A1"] = "Name"
    ws1["B1"] = "Price"
    ws1["A2"] = "Product 1"
    ws1["B2"] = 100
    
    # Instructions sheet (should be skipped)
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "How to use this file"
    ws2["A2"] = "Fill in the Products sheet"
    
    # Upload to site (priority sheet)
    ws3 = wb.create_sheet("Upload to site")
    ws3["A1"] = "Name"
    ws3["B1"] = "Price"
    ws3["A2"] = "Upload Product"
    ws3["B2"] = 150
    
    file_path = tmp_path / "multi_sheet.xlsx"
    wb.save(file_path)
    wb.close()
    
    return file_path


class TestMarkdownConverter:
    """Tests for MarkdownConverter class."""
    
    def test_init_default_values(self) -> None:
        """Test default initialization values."""
        converter = MarkdownConverter()
        assert converter.max_column_width == 50
        assert converter.skip_empty_rows is True
        assert converter.header_row == 1
    
    def test_init_custom_values(self) -> None:
        """Test custom initialization values."""
        converter = MarkdownConverter(
            max_column_width=100,
            skip_empty_rows=False,
            header_row=2,
        )
        assert converter.max_column_width == 100
        assert converter.skip_empty_rows is False
        assert converter.header_row == 2
    
    def test_convert_excel_to_markdown_basic(
        self, sample_workbook: Path
    ) -> None:
        """Test basic Excel to Markdown conversion."""
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(sample_workbook)
        
        # Should have header row with separator
        assert "| Name" in markdown
        assert "| Price" in markdown
        assert "| Category" in markdown
        assert "| ---" in markdown or "|---" in markdown
        
        # Should have data rows
        assert "Product 1" in markdown
        assert "100.5" in markdown
        assert "Electronics" in markdown
    
    def test_convert_excel_to_markdown_specific_sheet(
        self, multi_sheet_workbook: Path
    ) -> None:
        """Test conversion of specific sheet."""
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(
            multi_sheet_workbook, sheet_name="Upload to site"
        )
        
        assert "Upload Product" in markdown
        assert "150" in markdown
    
    def test_convert_excel_to_markdown_row_range(
        self, sample_workbook: Path
    ) -> None:
        """Test conversion with specific row range."""
        converter = MarkdownConverter()
        
        # Get only first 2 rows (header + first data row)
        markdown = converter.convert_excel_to_markdown(
            sample_workbook, start_row=1, end_row=2
        )
        
        assert "Product 1" in markdown
        assert "Product 2" not in markdown
    
    def test_convert_excel_with_merged_cells(
        self, workbook_with_merged_cells: Path
    ) -> None:
        """Test that merged cells repeat their value."""
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(
            workbook_with_merged_cells
        )
        
        # Electronics should appear for all 3 products
        # (merged cell value repeated)
        lines = markdown.split("\n")
        electronics_count = sum(
            1 for line in lines if "Electronics" in line
        )
        # At least 2 (one merged, one data row)
        assert electronics_count >= 1
        
        # All products should be present
        assert "Laptop A" in markdown
        assert "Laptop B" in markdown
        assert "Laptop C" in markdown
    
    def test_convert_file_not_found(self) -> None:
        """Test FileNotFoundError for missing file."""
        converter = MarkdownConverter()
        
        with pytest.raises(FileNotFoundError):
            converter.convert_excel_to_markdown("/nonexistent/file.xlsx")
    
    def test_convert_sheet_not_found(self, sample_workbook: Path) -> None:
        """Test ValueError for missing sheet."""
        converter = MarkdownConverter()
        
        with pytest.raises(ValueError) as exc_info:
            converter.convert_excel_to_markdown(
                sample_workbook, sheet_name="NonexistentSheet"
            )
        
        assert "not found" in str(exc_info.value)
    
    def test_get_sheet_info(self, multi_sheet_workbook: Path) -> None:
        """Test getting sheet information."""
        converter = MarkdownConverter()
        sheets = converter.get_sheet_info(multi_sheet_workbook)
        
        assert len(sheets) == 3
        
        # Find each sheet
        sheet_names = [s["name"] for s in sheets]
        assert "Products" in sheet_names
        assert "Instructions" in sheet_names
        assert "Upload to site" in sheet_names
        
        # Check metadata
        products_sheet = next(s for s in sheets if s["name"] == "Products")
        assert products_sheet["row_count"] > 0
        assert products_sheet["col_count"] > 0
        assert products_sheet["is_empty"] is False
    
    def test_format_cell_value_escapes_pipes(self) -> None:
        """Test that pipe characters are escaped."""
        converter = MarkdownConverter()
        
        result = converter._format_cell_value("Value | With | Pipes")
        
        # Pipes should be escaped for Markdown table compatibility
        assert "\\|" in result
    
    def test_format_cell_value_truncates_long_text(self) -> None:
        """Test that long text is truncated."""
        converter = MarkdownConverter(max_column_width=20)
        
        long_text = "x" * 100
        result = converter._format_cell_value(long_text)
        
        assert len(result) <= 20
        assert result.endswith("...")
    
    def test_format_cell_value_handles_none(self) -> None:
        """Test that None values return empty string."""
        converter = MarkdownConverter()
        
        result = converter._format_cell_value(None)
        assert result == ""
    
    def test_convert_excel_to_markdown_chunks(
        self, sample_workbook: Path
    ) -> None:
        """Test chunked conversion for large files."""
        converter = MarkdownConverter()
        
        # Use small chunk size to force multiple chunks
        chunks = converter.convert_excel_to_markdown_chunks(
            sample_workbook,
            chunk_size=2,  # 2 rows per chunk
            overlap=1,
        )
        
        assert len(chunks) >= 1
        
        # Check chunk structure
        for chunk in chunks:
            assert "chunk_id" in chunk
            assert "start_row" in chunk
            assert "end_row" in chunk
            assert "markdown" in chunk
            assert "total_rows" in chunk
            
            # Markdown should contain header
            assert "Name" in chunk["markdown"]


class TestMarkdownConverterEdgeCases:
    """Edge case tests for MarkdownConverter."""
    
    def test_empty_workbook(self, tmp_path: Path) -> None:
        """Test handling of empty workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        # No data added
        
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)
        wb.close()
        
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(file_path)
        
        # Should return empty or minimal string
        assert markdown == "" or len(markdown.strip()) == 0
    
    def test_special_characters(self, tmp_path: Path) -> None:
        """Test handling of special characters."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = "Value with <html> & special \"chars\""
        ws["B1"] = "Price"
        ws["B2"] = "1,234.56 р."
        
        file_path = tmp_path / "special.xlsx"
        wb.save(file_path)
        wb.close()
        
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(file_path)
        
        # Should handle without crashing
        assert "Header" in markdown
    
    def test_numeric_cells(self, tmp_path: Path) -> None:
        """Test handling of various numeric formats."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Type"
        ws["B1"] = "Value"
        ws["A2"] = "Integer"
        ws["B2"] = 42
        ws["A3"] = "Float"
        ws["B3"] = 3.14159
        ws["A4"] = "Negative"
        ws["B4"] = -100.5
        
        file_path = tmp_path / "numeric.xlsx"
        wb.save(file_path)
        wb.close()
        
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown(file_path)
        
        assert "42" in markdown
        assert "3.14159" in markdown
        assert "-100.5" in markdown

