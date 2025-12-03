"""
Unit Tests for ExcelStrategy
=============================

Tests for Excel file parsing with merged cell support.
"""

from decimal import Decimal
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest

from src.ingest.excel_strategy import ExcelStrategy
from src.schemas.domain import NormalizedRow


@pytest.fixture
def excel_strategy() -> ExcelStrategy:
    """Create default ExcelStrategy instance."""
    return ExcelStrategy()


@pytest.fixture
def sample_excel_path(tmp_path: Path) -> Path:
    """Create a sample Excel file for testing."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Header row
        ws["A1"] = "Название"
        ws["B1"] = "Цена"
        ws["C1"] = "Артикул"
        ws["D1"] = "Категория"

        # Data rows
        ws["A2"] = "Товар 1"
        ws["B2"] = 100.50
        ws["C2"] = "SKU-001"
        ws["D2"] = "Категория A"

        ws["A3"] = "Товар 2"
        ws["B3"] = 200.00
        ws["C3"] = "SKU-002"
        ws["D3"] = "Категория B"

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path
    except ImportError:
        pytest.skip("openpyxl not installed")


@pytest.fixture
def merged_cell_excel_path(tmp_path: Path) -> Path:
    """Create Excel file with merged cells for testing forward-fill."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Header row
        ws["A1"] = "Категория"
        ws["B1"] = "Название"
        ws["C1"] = "Цена"

        # Category merged across 3 rows
        ws["A2"] = "Электроника"
        ws.merge_cells("A2:A4")

        ws["B2"] = "Телефон"
        ws["C2"] = 1000

        ws["B3"] = "Планшет"
        ws["C3"] = 2000

        ws["B4"] = "Ноутбук"
        ws["C4"] = 3000

        file_path = tmp_path / "merged.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path
    except ImportError:
        pytest.skip("openpyxl not installed")


class TestExcelStrategySupportedExtensions:
    """Tests for supported_extensions property."""

    def test_supports_xlsx(self, excel_strategy: ExcelStrategy) -> None:
        """Should support .xlsx files."""
        assert "xlsx" in excel_strategy.supported_extensions

    def test_supports_xls(self, excel_strategy: ExcelStrategy) -> None:
        """Should support .xls files."""
        assert "xls" in excel_strategy.supported_extensions

    def test_supports_xlsm(self, excel_strategy: ExcelStrategy) -> None:
        """Should support .xlsm files."""
        assert "xlsm" in excel_strategy.supported_extensions


class TestExcelStrategyCanHandle:
    """Tests for can_handle method."""

    def test_can_handle_xlsx(self, excel_strategy: ExcelStrategy) -> None:
        """Should handle .xlsx files."""
        assert excel_strategy.can_handle("file.xlsx")

    def test_can_handle_xls(self, excel_strategy: ExcelStrategy) -> None:
        """Should handle .xls files."""
        assert excel_strategy.can_handle("file.xls")

    def test_cannot_handle_pdf(self, excel_strategy: ExcelStrategy) -> None:
        """Should not handle .pdf files."""
        assert not excel_strategy.can_handle("file.pdf")

    def test_cannot_handle_csv(self, excel_strategy: ExcelStrategy) -> None:
        """Should not handle .csv files."""
        assert not excel_strategy.can_handle("file.csv")


class TestExcelStrategyParse:
    """Tests for parse method."""

    @pytest.mark.asyncio
    async def test_parse_returns_normalized_rows(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Parsing should return list of NormalizedRow."""
        rows = await excel_strategy.parse(sample_excel_path)

        assert isinstance(rows, list)
        assert len(rows) == 2
        assert all(isinstance(r, NormalizedRow) for r in rows)

    @pytest.mark.asyncio
    async def test_parse_extracts_name(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Parsing should extract product names."""
        rows = await excel_strategy.parse(sample_excel_path)

        assert rows[0].name == "Товар 1"
        assert rows[1].name == "Товар 2"

    @pytest.mark.asyncio
    async def test_parse_extracts_price(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Parsing should extract prices as Decimal."""
        rows = await excel_strategy.parse(sample_excel_path)

        assert rows[0].price == Decimal("100.5")
        assert rows[1].price == Decimal("200")

    @pytest.mark.asyncio
    async def test_parse_extracts_sku(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Parsing should extract SKUs."""
        rows = await excel_strategy.parse(sample_excel_path)

        assert rows[0].sku == "SKU-001"
        assert rows[1].sku == "SKU-002"

    @pytest.mark.asyncio
    async def test_parse_extracts_category(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Parsing should extract categories."""
        rows = await excel_strategy.parse(sample_excel_path)

        assert rows[0].category == "Категория A"
        assert rows[1].category == "Категория B"

    @pytest.mark.asyncio
    async def test_parse_nonexistent_file_raises(
        self,
        excel_strategy: ExcelStrategy,
    ) -> None:
        """Parsing nonexistent file should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            await excel_strategy.parse("/nonexistent/file.xlsx")


class TestExcelStrategyMergedCells:
    """Tests for merged cell forward-fill functionality."""

    @pytest.mark.asyncio
    async def test_forward_fill_merged_category(
        self,
        excel_strategy: ExcelStrategy,
        merged_cell_excel_path: Path,
    ) -> None:
        """Merged cells should be forward-filled to all rows."""
        rows = await excel_strategy.parse(merged_cell_excel_path)

        # All 3 rows should have the category from the merged cell
        assert len(rows) == 3
        assert all(r.category == "Электроника" for r in rows)

    @pytest.mark.asyncio
    async def test_forward_fill_preserves_other_columns(
        self,
        excel_strategy: ExcelStrategy,
        merged_cell_excel_path: Path,
    ) -> None:
        """Forward-fill should not affect non-merged columns."""
        rows = await excel_strategy.parse(merged_cell_excel_path)

        # Names should be unique
        names = [r.name for r in rows]
        assert names == ["Телефон", "Планшет", "Ноутбук"]

        # Prices should be unique
        prices = [r.price for r in rows]
        assert prices == [Decimal("1000"), Decimal("2000"), Decimal("3000")]


class TestExcelStrategyValidation:
    """Tests for validate_file method."""

    @pytest.mark.asyncio
    async def test_validate_existing_file(
        self,
        excel_strategy: ExcelStrategy,
        sample_excel_path: Path,
    ) -> None:
        """Valid Excel file should pass validation."""
        is_valid = await excel_strategy.validate_file(sample_excel_path)
        assert is_valid is True

    @pytest.mark.asyncio
    async def test_validate_nonexistent_file(
        self,
        excel_strategy: ExcelStrategy,
    ) -> None:
        """Nonexistent file should fail validation."""
        is_valid = await excel_strategy.validate_file("/nonexistent/file.xlsx")
        assert is_valid is False

    @pytest.mark.asyncio
    async def test_validate_wrong_extension(
        self,
        excel_strategy: ExcelStrategy,
        tmp_path: Path,
    ) -> None:
        """File with wrong extension should fail validation."""
        pdf_file = tmp_path / "test.pdf"
        pdf_file.write_text("not a pdf")

        is_valid = await excel_strategy.validate_file(pdf_file)
        assert is_valid is False


class TestExcelStrategyConfiguration:
    """Tests for strategy configuration options."""

    def test_custom_header_row(self) -> None:
        """Custom header row should be respected."""
        strategy = ExcelStrategy(header_row=2)
        assert strategy._header_row == 2

    def test_custom_skip_rows(self) -> None:
        """Custom skip_rows should be respected."""
        strategy = ExcelStrategy(skip_rows=5)
        assert strategy._skip_rows == 5

    def test_custom_max_rows(self) -> None:
        """Custom max_rows should be respected."""
        strategy = ExcelStrategy(max_rows=100)
        assert strategy._max_rows == 100

    def test_custom_sheet_name(self) -> None:
        """Custom sheet_name should be respected."""
        strategy = ExcelStrategy(sheet_name="Data")
        assert strategy._sheet_name == "Data"


class TestExcelStrategyColumnDetection:
    """Tests for automatic column detection."""

    def test_detect_russian_name_column(
        self,
        excel_strategy: ExcelStrategy,
    ) -> None:
        """Should detect Russian 'название' as name column."""
        columns = ["Артикул", "Название", "Цена"]
        mapping = excel_strategy._detect_column_mapping(columns)

        assert mapping.get("name") == "Название"

    def test_detect_russian_price_column(
        self,
        excel_strategy: ExcelStrategy,
    ) -> None:
        """Should detect Russian 'цена' as price column."""
        columns = ["Название", "Цена", "Код"]
        mapping = excel_strategy._detect_column_mapping(columns)

        assert mapping.get("price") == "Цена"

    def test_detect_english_columns(
        self,
        excel_strategy: ExcelStrategy,
    ) -> None:
        """Should detect English column names."""
        columns = ["Product Name", "Price", "SKU"]
        mapping = excel_strategy._detect_column_mapping(columns)

        assert mapping.get("name") == "Product Name"
        assert mapping.get("price") == "Price"
        assert mapping.get("sku") == "SKU"

