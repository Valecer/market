"""
Ingestion Service
==================

Orchestrates the file parsing pipeline:
1. Select appropriate parser based on file type
2. Parse file to normalized rows (traditional or ML-based)
3. Insert items into supplier_items table
4. Log any errors to parsing_logs table
5. Return chunk data for embedding generation

Supports two parsing modes:
- Traditional: Uses ParserFactory with rule-based extraction
- ML-based: Uses TwoStageParsingService with LLM extraction (Phase 10)

Follows Single Responsibility: Only handles ingestion orchestration.
"""

from pathlib import Path
from typing import Any
from uuid import UUID

from src.config.settings import get_settings
from src.db.connection import get_session
from src.db.repositories.parsing_logs_repo import ParsingLogsRepository
from src.db.repositories.supplier_items_repo import SupplierItemsRepository
from src.ingest.chunker import Chunker
from src.ingest.parser_factory import ParserFactory
from src.ingest.table_normalizer import ParseResult, ParserError
from src.schemas.domain import ChunkData, NormalizedRow, ParsingMetrics
from src.services.two_stage_parser import TwoStageParsingService
from src.utils.errors import ParsingError
from src.utils.logger import get_logger

logger = get_logger(__name__)


class IngestionResult:
    """
    Result of an ingestion operation.

    Attributes:
        success: Whether ingestion completed successfully
        total_rows: Total rows found in file
        processed_rows: Rows successfully processed
        error_count: Number of errors encountered
        item_ids: List of created supplier_item IDs
        chunks: List of ChunkData for embedding generation
        errors: List of error details
        metrics: Parsing metrics from ML-based parsing (Phase 10)
        parsing_mode: Mode used for parsing ('traditional' or 'ml')
    """

    def __init__(
        self,
        success: bool = True,
        total_rows: int = 0,
        processed_rows: int = 0,
        error_count: int = 0,
        item_ids: list[UUID] | None = None,
        chunks: list[ChunkData] | None = None,
        errors: list[dict[str, Any]] | None = None,
        metrics: ParsingMetrics | None = None,
        parsing_mode: str = "traditional",
    ) -> None:
        self.success = success
        self.total_rows = total_rows
        self.processed_rows = processed_rows
        self.error_count = error_count
        self.item_ids = item_ids or []
        self.chunks = chunks or []
        self.errors = errors or []
        self.metrics = metrics
        self.parsing_mode = parsing_mode

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        result = {
            "success": self.success,
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "error_count": self.error_count,
            "item_count": len(self.item_ids),
            "chunk_count": len(self.chunks),
            "parsing_mode": self.parsing_mode,
        }
        if self.metrics:
            result["metrics"] = self.metrics.model_dump()
        return result


class IngestionService:
    """
    Service for orchestrating file ingestion.

    Supports two parsing modes:
    - Traditional: Uses ParserFactory with rule-based extraction
    - ML-based: Uses TwoStageParsingService with LLM extraction (Phase 10)

    Workflow:
        1. Validate file exists and type is supported
        2. Create parser strategy based on file type and mode
        3. Parse file to NormalizedRow objects
        4. Insert rows into supplier_items table
        5. Generate chunks for embedding
        6. Log any errors encountered

    Usage:
        service = IngestionService()
        result = await service.ingest_file(
            file_path="/path/to/file.xlsx",
            supplier_id=uuid,
            job_id=uuid,
            use_ml_parsing=True,  # Enable ML-based parsing
        )

        # Check result
        if result.success:
            print(f"Processed {result.processed_rows} rows")
            print(f"Metrics: {result.metrics}")
            for chunk in result.chunks:
                # Generate embeddings...
    """

    def __init__(
        self,
        chunker: Chunker | None = None,
        two_stage_parser: TwoStageParsingService | None = None,
    ) -> None:
        """
        Initialize ingestion service.

        Args:
            chunker: Optional custom Chunker instance
            two_stage_parser: Optional TwoStageParsingService for ML parsing
        """
        self._chunker = chunker or Chunker()
        self._two_stage_parser = two_stage_parser
        self._settings = get_settings()
        logger.debug("IngestionService initialized")

    async def ingest_file(
        self,
        file_path: str | Path,
        supplier_id: UUID,
        job_id: UUID | None = None,
        file_type: str | None = None,
        parser_kwargs: dict[str, Any] | None = None,
        use_ml_parsing: bool = False,
        default_currency: str | None = None,
        composite_delimiter: str = "|",
    ) -> IngestionResult:
        """
        Ingest a file and create supplier items.

        Supports two parsing modes:
        - Traditional (use_ml_parsing=False): Uses ParserFactory
        - ML-based (use_ml_parsing=True): Uses TwoStageParsingService

        Args:
            file_path: Path to file to ingest
            supplier_id: Supplier ID for created items
            job_id: Optional job ID for tracking
            file_type: Optional explicit file type (auto-detected if None)
            parser_kwargs: Optional arguments for parser
            use_ml_parsing: Enable ML-based parsing with TwoStageParsingService
            default_currency: Default ISO 4217 currency code for ML parsing
            composite_delimiter: Delimiter for composite names in ML parsing

        Returns:
            IngestionResult with processing details and metrics

        Raises:
            ParsingError: If file cannot be parsed
            FileNotFoundError: If file doesn't exist
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path

        logger.info(
            "Starting file ingestion",
            file_path=str(path),
            supplier_id=str(supplier_id),
            job_id=str(job_id) if job_id else None,
            use_ml_parsing=use_ml_parsing,
        )

        # Validate file exists
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        # Route to appropriate parsing method
        if use_ml_parsing:
            return await self._ingest_with_ml_parsing(
                path=path,
                supplier_id=supplier_id,
                job_id=job_id,
                file_type=file_type,
                default_currency=default_currency,
                composite_delimiter=composite_delimiter,
            )
        else:
            return await self._ingest_with_traditional_parsing(
                path=path,
                supplier_id=supplier_id,
                job_id=job_id,
                file_type=file_type,
                parser_kwargs=parser_kwargs,
            )

    async def _ingest_with_traditional_parsing(
        self,
        path: Path,
        supplier_id: UUID,
        job_id: UUID | None,
        file_type: str | None,
        parser_kwargs: dict[str, Any] | None,
    ) -> IngestionResult:
        """
        Traditional parsing using ParserFactory.

        Uses rule-based extraction strategies (Excel, PDF).
        """
        errors: list[dict[str, Any]] = []

        try:
            # Create parser
            kwargs = parser_kwargs or {}
            if file_type:
                parser = ParserFactory.create(file_type, **kwargs)
            else:
                parser = ParserFactory.from_file_path(path, **kwargs)

            # Parse file
            logger.debug("Parsing file (traditional)", parser=type(parser).__name__)
            rows = await parser.parse(path, supplier_id)

            if not rows:
                logger.warning("No rows parsed from file", file_path=str(path))
                return IngestionResult(
                    success=True,
                    total_rows=0,
                    processed_rows=0,
                    parsing_mode="traditional",
                )

            # Insert into database
            async with get_session() as session:
                items_repo = SupplierItemsRepository(session)
                logs_repo = ParsingLogsRepository(session)

                # Create supplier items
                item_ids = await items_repo.create_batch(
                    supplier_id=supplier_id,
                    rows=rows,
                    source_type="traditional",
                )

                # Generate chunks for embedding
                chunks = self._chunker.chunk_batch(rows, item_ids)

                # Log any errors
                if errors:
                    await logs_repo.log_batch(
                        supplier_id=supplier_id,
                        errors=errors,
                        job_id=job_id,
                    )

            logger.info(
                "Traditional parsing complete",
                file_path=str(path),
                total_rows=len(rows),
                items_created=len(item_ids),
                chunks_generated=len(chunks),
            )

            return IngestionResult(
                success=True,
                total_rows=len(rows),
                processed_rows=len(item_ids),
                error_count=len(errors),
                item_ids=item_ids,
                chunks=chunks,
                errors=errors,
                parsing_mode="traditional",
            )

        except ParsingError as e:
            logger.error(
                "Parsing error during traditional ingestion",
                file_path=str(path),
                error=e.message,
            )
            await self._log_parsing_error(supplier_id, job_id, e)
            return IngestionResult(
                success=False,
                error_count=1,
                errors=[{"type": "parsing", "message": e.message, "details": e.details}],
                parsing_mode="traditional",
            )

        except Exception as e:
            logger.exception(
                "Unexpected error during traditional ingestion",
                file_path=str(path),
                error=str(e),
            )
            return IngestionResult(
                success=False,
                error_count=1,
                errors=[{"type": "unknown", "message": str(e)}],
                parsing_mode="traditional",
            )

    async def _ingest_with_ml_parsing(
        self,
        path: Path,
        supplier_id: UUID,
        job_id: UUID | None,
        file_type: str | None,
        default_currency: str | None,
        composite_delimiter: str,
    ) -> IngestionResult:
        """
        ML-based parsing using TwoStageParsingService.

        Uses two-stage LLM extraction with structure analysis.
        Phase 10 enhancement.
        """
        errors: list[dict[str, Any]] = []

        try:
            # Get or create TwoStageParsingService
            parser_service = self._two_stage_parser or TwoStageParsingService()

            # Read raw data from file using traditional parser
            # (TwoStageParser works on raw table data, not files directly)
            raw_data = await self._read_raw_table_data(path, file_type)

            if not raw_data:
                logger.warning("No data read from file", file_path=str(path))
                return IngestionResult(
                    success=True,
                    total_rows=0,
                    processed_rows=0,
                    parsing_mode="ml",
                )

            logger.debug(
                "ML parsing file",
                file_path=str(path),
                raw_rows=len(raw_data),
            )

            # Parse using two-stage ML service
            rows, metrics = await parser_service.parse_document(
                raw_data=raw_data,
                supplier_id=supplier_id,
                default_currency=default_currency,
                composite_delimiter=composite_delimiter,
            )

            if not rows:
                logger.warning("No rows parsed from file (ML)", file_path=str(path))
                return IngestionResult(
                    success=True,
                    total_rows=metrics.total_rows,
                    processed_rows=0,
                    metrics=metrics,
                    parsing_mode="ml",
                )

            # Insert into database
            async with get_session() as session:
                items_repo = SupplierItemsRepository(session)
                logs_repo = ParsingLogsRepository(session)

                # Create supplier items
                item_ids = await items_repo.create_batch(
                    supplier_id=supplier_id,
                    rows=rows,
                    source_type="ml_analyzed",
                )

                # Generate chunks for embedding
                chunks = self._chunker.chunk_batch(rows, item_ids)

                # Log any errors
                if errors:
                    await logs_repo.log_batch(
                        supplier_id=supplier_id,
                        errors=errors,
                        job_id=job_id,
                    )

            logger.info(
                "ML parsing complete",
                file_path=str(path),
                total_rows=metrics.total_rows,
                parsed_rows=metrics.parsed_rows,
                items_created=len(item_ids),
                chunks_generated=len(chunks),
                total_tokens=metrics.total_tokens,
                duration_ms=metrics.duration_ms,
            )

            return IngestionResult(
                success=True,
                total_rows=metrics.total_rows,
                processed_rows=len(item_ids),
                error_count=metrics.error_rows,
                item_ids=item_ids,
                chunks=chunks,
                errors=errors,
                metrics=metrics,
                parsing_mode="ml",
            )

        except ParsingError as e:
            logger.error(
                "Parsing error during ML ingestion",
                file_path=str(path),
                error=e.message,
            )
            await self._log_parsing_error(supplier_id, job_id, e)
            return IngestionResult(
                success=False,
                error_count=1,
                errors=[{"type": "parsing", "message": e.message, "details": e.details}],
                parsing_mode="ml",
            )

        except Exception as e:
            logger.exception(
                "Unexpected error during ML ingestion",
                file_path=str(path),
                error=str(e),
            )
            return IngestionResult(
                success=False,
                error_count=1,
                errors=[{"type": "unknown", "message": str(e)}],
                parsing_mode="ml",
            )

    async def _read_raw_table_data(
        self,
        path: Path,
        file_type: str | None,
    ) -> list[list[str]]:
        """
        Read raw table data from file.

        Uses openpyxl directly for Excel files to get raw cell values.

        Args:
            path: Path to file
            file_type: File type hint

        Returns:
            List of rows, each row is list of string values
        """
        from openpyxl import load_workbook

        extension = file_type or path.suffix.lower().lstrip(".")

        if extension in ("xlsx", "xls", "xlsm", "excel"):
            wb = load_workbook(str(path), data_only=True)
            ws = wb.active or wb.worksheets[0]

            raw_data: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(v.strip() for v in row_values):
                    raw_data.append(row_values)

            wb.close()
            return raw_data

        elif extension == "pdf":
            # For PDF, use pymupdf4llm to extract text
            # Then parse into table-like structure
            import pymupdf4llm

            md_text = pymupdf4llm.to_markdown(str(path))
            # Parse markdown tables into raw data
            lines = md_text.split("\n")
            raw_data = []
            for line in lines:
                if "|" in line and not line.strip().startswith("---"):
                    cells = [c.strip() for c in line.split("|") if c.strip()]
                    if cells:
                        raw_data.append(cells)
            return raw_data

        else:
            # Fallback: try to read as CSV-like
            raw_data = []
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    cells = line.strip().split(",")
                    if any(c.strip() for c in cells):
                        raw_data.append(cells)
            return raw_data

    async def _log_parsing_error(
        self,
        supplier_id: UUID,
        job_id: UUID | None,
        error: ParsingError,
    ) -> None:
        """Log a parsing error to the database."""
        try:
            async with get_session() as session:
                logs_repo = ParsingLogsRepository(session)
                await logs_repo.log_error(
                    supplier_id=supplier_id,
                    error_type="parsing",
                    message=error.message,
                    severity="error",
                    job_id=job_id,
                    context=error.details,
                )
        except Exception as log_error:
            logger.warning("Failed to log error", error=str(log_error))

    async def parse_only(
        self,
        file_path: str | Path,
        file_type: str | None = None,
        parser_kwargs: dict[str, Any] | None = None,
    ) -> list[NormalizedRow]:
        """
        Parse a file without inserting into database.

        Useful for previewing file contents or testing parsers.

        Args:
            file_path: Path to file
            file_type: Optional explicit file type
            parser_kwargs: Optional arguments for parser

        Returns:
            List of NormalizedRow objects
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        kwargs = parser_kwargs or {}
        if file_type:
            parser = ParserFactory.create(file_type, **kwargs)
        else:
            parser = ParserFactory.from_file_path(path, **kwargs)

        return await parser.parse(path)

    async def validate_file(
        self,
        file_path: str | Path,
        file_type: str | None = None,
    ) -> bool:
        """
        Validate that a file can be parsed.

        Args:
            file_path: Path to file
            file_type: Optional explicit file type

        Returns:
            True if file can be parsed
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path

        if not path.exists():
            return False

        try:
            if file_type:
                parser = ParserFactory.create(file_type)
            else:
                parser = ParserFactory.from_file_path(path)
            return await parser.validate_file(path)
        except Exception:
            return False


# Convenience function for simple ingestion
async def ingest_file(
    file_path: str | Path,
    supplier_id: UUID,
    job_id: UUID | None = None,
    use_ml_parsing: bool = False,
    default_currency: str | None = None,
    composite_delimiter: str = "|",
    **kwargs: Any,
) -> IngestionResult:
    """
    Convenience function to ingest a file.

    Args:
        file_path: Path to file
        supplier_id: Supplier ID
        job_id: Optional job ID
        use_ml_parsing: Enable ML-based parsing
        default_currency: Default currency for ML parsing
        composite_delimiter: Delimiter for composite names
        **kwargs: Additional arguments for parser

    Returns:
        IngestionResult
    """
    service = IngestionService()
    return await service.ingest_file(
        file_path=file_path,
        supplier_id=supplier_id,
        job_id=job_id,
        use_ml_parsing=use_ml_parsing,
        default_currency=default_currency,
        composite_delimiter=composite_delimiter,
        parser_kwargs=kwargs,
    )

