"""
Markdown Converter for Semantic ETL Pipeline
============================================

Converts Excel/CSV files to Markdown table format for LLM processing.
Preserves visual layout including merged cells, column alignment,
and spatial relationships.

Phase 9: Semantic ETL Pipeline Refactoring
"""

import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class MarkdownConverter:
    """
    Converts Excel worksheets to Markdown table format.
    
    Features:
    - Merged cell handling (repeats value across merged range)
    - Configurable row limits for chunking
    - Header row detection
    - Empty row/column trimming
    
    Example:
        converter = MarkdownConverter()
        markdown = converter.convert_excel_to_markdown("file.xlsx", "Sheet1")
        
        # Or convert specific rows for chunking
        chunks = converter.convert_excel_to_markdown_chunks(
            "file.xlsx", "Sheet1", chunk_size=250, overlap=40
        )
    """
    
    def __init__(
        self,
        max_column_width: int = 50,
        skip_empty_rows: bool = True,
        header_row: int = 1,
    ):
        """
        Initialize MarkdownConverter.
        
        Args:
            max_column_width: Maximum characters per column (truncated with ...)
            skip_empty_rows: Whether to skip completely empty rows
            header_row: Row number containing headers (1-indexed)
        """
        self.max_column_width = max_column_width
        self.skip_empty_rows = skip_empty_rows
        self.header_row = header_row
    
    def convert_excel_to_markdown(
        self,
        file_path: str | Path,
        sheet_name: Optional[str] = None,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> str:
        """
        Convert Excel worksheet to Markdown table.
        
        Args:
            file_path: Path to Excel file (.xlsx, .xlsm)
            sheet_name: Sheet name to convert (None = active sheet)
            start_row: Starting row number (1-indexed, None = from beginning)
            end_row: Ending row number (1-indexed, inclusive, None = to end)
        
        Returns:
            Markdown table string with pipe-delimited columns
        
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If sheet_name doesn't exist
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available: {workbook.sheetnames}"
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            if worksheet is None:
                raise ValueError("No active worksheet found")
            
            return self._worksheet_to_markdown(worksheet, start_row, end_row)
        finally:
            workbook.close()
    
    def _worksheet_to_markdown(
        self,
        worksheet: Worksheet,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> str:
        """
        Convert worksheet to Markdown table.
        
        Args:
            worksheet: openpyxl Worksheet object
            start_row: Starting row (1-indexed)
            end_row: Ending row (1-indexed, inclusive)
        
        Returns:
            Markdown table string
        """
        # Get merged cell ranges
        merged_ranges = self._build_merged_cell_map(worksheet)
        
        # Determine actual data bounds
        min_row = start_row or 1
        max_row = end_row or worksheet.max_row or 1
        min_col = 1
        max_col = worksheet.max_column or 1
        
        # Handle edge case of empty worksheet
        if max_row < min_row or max_col < min_col:
            return ""
        
        # Extract data with merged cell handling
        rows_data: list[list[str]] = []
        
        for row_idx in range(min_row, max_row + 1):
            row_values: list[str] = []
            is_empty_row = True
            
            for col_idx in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Check if this cell is part of a merged range
                if isinstance(cell, MergedCell):
                    # Get value from the top-left cell of the merge range
                    value = self._get_merged_cell_value(
                        worksheet, row_idx, col_idx, merged_ranges
                    )
                else:
                    value = cell.value
                
                # Format value
                formatted = self._format_cell_value(value)
                row_values.append(formatted)
                
                if formatted.strip():
                    is_empty_row = False
            
            # Skip empty rows if configured
            if self.skip_empty_rows and is_empty_row:
                continue
            
            rows_data.append(row_values)
        
        if not rows_data:
            return ""
        
        # Calculate column widths for alignment
        col_widths = self._calculate_column_widths(rows_data)
        
        # Build Markdown table
        lines: list[str] = []
        
        for i, row in enumerate(rows_data):
            # Format cells with padding
            cells = [
                self._pad_cell(cell, col_widths[j])
                for j, cell in enumerate(row)
            ]
            lines.append(f"| {' | '.join(cells)} |")
            
            # Add separator after header row
            if i == 0:
                separators = ["-" * w for w in col_widths]
                lines.append(f"| {' | '.join(separators)} |")
        
        return "\n".join(lines)
    
    def _build_merged_cell_map(
        self, worksheet: Worksheet
    ) -> dict[tuple[int, int], tuple[int, int]]:
        """
        Build a map from merged cell coordinates to their top-left cell.
        
        Args:
            worksheet: openpyxl Worksheet
        
        Returns:
            Dict mapping (row, col) -> (top_left_row, top_left_col)
        """
        merged_map: dict[tuple[int, int], tuple[int, int]] = {}
        
        for merge_range in worksheet.merged_cells.ranges:
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merged_map[(row, col)] = (min_row, min_col)
        
        return merged_map
    
    def _get_merged_cell_value(
        self,
        worksheet: Worksheet,
        row: int,
        col: int,
        merged_map: dict[tuple[int, int], tuple[int, int]],
    ) -> Optional[str]:
        """
        Get value for a merged cell by looking up its top-left origin.
        
        Args:
            worksheet: openpyxl Worksheet
            row: Current row
            col: Current column
            merged_map: Map from cell coords to top-left coords
        
        Returns:
            Cell value from the merge origin
        """
        if (row, col) in merged_map:
            origin_row, origin_col = merged_map[(row, col)]
            return worksheet.cell(row=origin_row, column=origin_col).value
        return None
    
    def _format_cell_value(self, value: object) -> str:
        """
        Format cell value as string with cleanup.
        
        Args:
            value: Raw cell value (any type)
        
        Returns:
            Formatted string value
        """
        if value is None:
            return ""
        
        # Convert to string
        text = str(value).strip()
        
        # Remove newlines (replace with space)
        text = text.replace("\n", " ").replace("\r", " ")
        
        # Collapse multiple spaces
        text = " ".join(text.split())
        
        # Escape pipe characters (Markdown table delimiter)
        text = text.replace("|", "\\|")
        
        # Truncate if too long
        if len(text) > self.max_column_width:
            text = text[: self.max_column_width - 3] + "..."
        
        return text
    
    def _calculate_column_widths(
        self, rows_data: list[list[str]]
    ) -> list[int]:
        """
        Calculate optimal column widths based on content.
        
        Args:
            rows_data: List of rows, each row is a list of cell strings
        
        Returns:
            List of column widths (minimum 3 for Markdown separator)
        """
        if not rows_data:
            return []
        
        num_cols = len(rows_data[0])
        widths = [3] * num_cols  # Minimum width for "---"
        
        for row in rows_data:
            for i, cell in enumerate(row):
                if i < num_cols:
                    widths[i] = max(widths[i], len(cell))
        
        # Cap at max_column_width
        return [min(w, self.max_column_width) for w in widths]
    
    def _pad_cell(self, value: str, width: int) -> str:
        """
        Pad cell value to specified width.
        
        Args:
            value: Cell value
            width: Target width
        
        Returns:
            Padded string
        """
        return value.ljust(width)
    
    def convert_excel_to_markdown_chunks(
        self,
        file_path: str | Path,
        sheet_name: Optional[str] = None,
        chunk_size: int = 250,
        overlap: int = 40,
    ) -> list[dict]:
        """
        Convert Excel to Markdown in overlapping chunks for LLM processing.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (None = active sheet)
            chunk_size: Number of data rows per chunk (excluding header)
            overlap: Number of overlapping rows between chunks
        
        Returns:
            List of chunk dicts with keys:
            - chunk_id: int (0-indexed)
            - start_row: int (1-indexed, data row, excluding header)
            - end_row: int (1-indexed, data row)
            - markdown: str (includes header)
            - total_rows: int (data rows in this chunk)
        
        Example:
            chunks = converter.convert_excel_to_markdown_chunks(
                "products.xlsx", chunk_size=250, overlap=40
            )
            for chunk in chunks:
                llm_response = extract_products(chunk["markdown"])
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available: {workbook.sheetnames}"
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            if worksheet is None:
                raise ValueError("No active worksheet found")
            
            # Get total rows (excluding header)
            total_data_rows = (worksheet.max_row or 1) - self.header_row
            if total_data_rows <= 0:
                return []
            
            # Extract header as separate markdown
            header_markdown = self._worksheet_to_markdown(
                worksheet,
                start_row=self.header_row,
                end_row=self.header_row,
            )
            
            chunks: list[dict] = []
            chunk_id = 0
            current_start = self.header_row + 1  # First data row
            
            while current_start <= (worksheet.max_row or 1):
                current_end = min(
                    current_start + chunk_size - 1,
                    worksheet.max_row or current_start
                )
                
                # Get data rows for this chunk
                data_markdown = self._worksheet_to_markdown(
                    worksheet,
                    start_row=current_start,
                    end_row=current_end,
                )
                
                # Combine header with data (header already has separator)
                if data_markdown:
                    # Remove header line from data and just keep data rows
                    data_lines = data_markdown.split("\n")
                    # Skip any duplicate header detection
                    combined_markdown = header_markdown
                    if data_lines:
                        # Data markdown doesn't include header, so just append
                        # Actually we need to add data without header/separator
                        # Let's reconstruct properly
                        combined_markdown = header_markdown + "\n" + "\n".join(
                            line for i, line in enumerate(data_lines)
                            if i >= 2 or "---" not in line  # Skip separator lines in data
                        )
                else:
                    combined_markdown = header_markdown
                
                chunks.append({
                    "chunk_id": chunk_id,
                    "start_row": current_start - self.header_row,  # 1-indexed data row
                    "end_row": current_end - self.header_row,  # 1-indexed data row
                    "markdown": combined_markdown,
                    "total_rows": current_end - current_start + 1,
                })
                
                chunk_id += 1
                
                # Move to next chunk with overlap
                current_start = current_end - overlap + 1
                
                # Prevent infinite loop if overlap >= chunk_size
                if current_start <= chunks[-1]["start_row"] + self.header_row:
                    current_start = current_end + 1
                
                # Break if we've processed all rows
                if current_end >= (worksheet.max_row or 1):
                    break
            
            return chunks
        
        finally:
            workbook.close()
    
    def get_sheet_info(
        self, file_path: str | Path
    ) -> list[dict]:
        """
        Get information about all sheets in an Excel file.
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            List of sheet info dicts with keys:
            - name: str (sheet name)
            - row_count: int (total rows)
            - col_count: int (total columns)
            - is_empty: bool
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        try:
            sheets_info: list[dict] = []
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                row_count = worksheet.max_row or 0
                col_count = worksheet.max_column or 0
                
                sheets_info.append({
                    "name": sheet_name,
                    "row_count": row_count,
                    "col_count": col_count,
                    "is_empty": row_count == 0 or col_count == 0,
                })
            
            return sheets_info
        
        finally:
            workbook.close()

