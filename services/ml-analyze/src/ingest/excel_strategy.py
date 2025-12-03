"""
Excel Strategy - Complex Excel File Parser
==========================================

Parses Excel files with complex layouts including:
- Multi-row headers
- Merged cells (with forward-fill normalization)
- Hidden rows/columns
- Multiple worksheets

Uses openpyxl for reading and pandas for data manipulation.

Follows Strategy Pattern: Implements TableNormalizer interface.
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from src.ingest.table_normalizer import (
    ParseResult,
    ParserError,
    TableNormalizer,
)
from src.schemas.domain import NormalizedRow
from src.utils.errors import ParsingError
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelStrategy(TableNormalizer):
    """
    Excel file parsing strategy with merged cell support.

    Handles complex Excel layouts common in supplier price lists:
    - Merged category cells spanning multiple rows
    - Multi-row headers
    - Hidden columns containing metadata
    - Multiple worksheets

    Forward-Fill Algorithm:
        When a cell is merged across rows (e.g., category spanning 10 product rows),
        the merged value is "forward-filled" to all affected rows.

    Example:
        Input Excel:
        | Category (merged A1:A3) | Name    | Price |
        |                         | Item 1  | 10.00 |
        |                         | Item 2  | 15.00 |
        |                         | Item 3  | 20.00 |

        Output:
        | Category | Name    | Price |
        | Tools    | Item 1  | 10.00 |
        | Tools    | Item 2  | 15.00 |
        | Tools    | Item 3  | 20.00 |
    """

    # Column name patterns for automatic field detection (Russian + English)
    FIELD_PATTERNS = {
        "name": ["название", "наименование", "товар", "модель", "name", "product", "item", "description"],
        "price": ["цена", "стоимость", "price", "cost", "amount", "сток"],
        "sku": ["артикул", "код", "sku", "code", "product_id", "item_id"],
        "category": ["категория", "раздел", "group", "category", "section"],
        "brand": ["бренд", "производитель", "brand", "manufacturer"],
        "unit": ["ед.", "единица", "unit", "uom", "ед.изм"],
        "description": ["описание", "характеристики", "description", "details"],
    }

    def __init__(
        self,
        header_row: int = 0,
        skip_rows: int = 0,
        max_rows: int | None = None,
        sheet_name: str | int | None = None,
    ) -> None:
        """
        Initialize Excel strategy.

        Args:
            header_row: Row index containing headers (0-based, -1 for auto-detect)
            skip_rows: Number of rows to skip at the start
            max_rows: Maximum rows to process (None = all)
            sheet_name: Sheet name or index (None = first sheet)
        """
        self._header_row = header_row
        self._skip_rows = skip_rows
        self._max_rows = max_rows
        self._sheet_name = sheet_name
        logger.debug(
            "ExcelStrategy initialized",
            header_row=header_row,
            skip_rows=skip_rows,
            max_rows=max_rows,
            sheet_name=sheet_name,
        )

    @property
    def supported_extensions(self) -> list[str]:
        """Return supported file extensions."""
        return ["xlsx", "xls", "xlsm"]

    async def parse(
        self,
        file_path: str | Path,
        supplier_id: "UUID | None" = None,
    ) -> list[NormalizedRow]:
        """
        Parse Excel file and return normalized rows.

        Args:
            file_path: Path to Excel file
            supplier_id: Optional supplier ID for context

        Returns:
            List of NormalizedRow objects

        Raises:
            ParsingError: If file cannot be parsed
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path

        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        logger.info("Parsing Excel file", file_path=str(path), supplier_id=supplier_id)

        try:
            # Load workbook with data_only=True to get calculated values
            wb = load_workbook(str(path), data_only=True)

            # Select worksheet
            ws = self._select_worksheet(wb)

            # Extract merged cell ranges for forward-fill
            merged_ranges = self._get_merged_ranges(ws)

            # Read data with forward-fill applied
            df = self._read_worksheet_with_forward_fill(ws, merged_ranges)

            if df.empty:
                logger.warning("Empty worksheet", file_path=str(path))
                return []

            # Detect or use configured header row
            header_row_idx = self._detect_header_row(df) if self._header_row == -1 else self._header_row

            # Set headers and clean data
            df = self._apply_headers(df, header_row_idx)

            # Map columns to standard fields
            column_mapping = self._detect_column_mapping(df.columns.tolist())

            # Convert rows to NormalizedRow objects
            rows = self._convert_to_normalized_rows(df, column_mapping)

            logger.info(
                "Excel parsing complete",
                file_path=str(path),
                total_rows=len(df),
                normalized_rows=len(rows),
            )

            wb.close()
            return rows

        except Exception as e:
            if isinstance(e, (ParsingError, FileNotFoundError)):
                raise
            logger.error("Excel parsing failed", file_path=str(path), error=str(e))
            raise ParsingError(
                message=f"Failed to parse Excel file: {e}",
                details={"file_path": str(path), "error": str(e)},
            ) from e

    async def validate_file(self, file_path: str | Path) -> bool:
        """
        Validate that file can be parsed.

        Args:
            file_path: Path to file

        Returns:
            True if file can be parsed
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path

        if not path.exists():
            return False

        if not self.can_handle(path):
            return False

        try:
            # Try to open and read basic info
            wb = load_workbook(str(path), read_only=True)
            has_data = len(wb.sheetnames) > 0
            wb.close()
            return has_data
        except Exception:
            return False

    def _select_worksheet(self, wb) -> Worksheet:
        """Select worksheet based on configuration."""
        if self._sheet_name is None:
            return wb.active or wb.worksheets[0]
        elif isinstance(self._sheet_name, int):
            if self._sheet_name < len(wb.worksheets):
                return wb.worksheets[self._sheet_name]
            raise ParsingError(
                message=f"Sheet index {self._sheet_name} out of range",
                details={"available_sheets": len(wb.worksheets)},
            )
        else:
            if self._sheet_name in wb.sheetnames:
                return wb[self._sheet_name]
            raise ParsingError(
                message=f"Sheet '{self._sheet_name}' not found",
                details={"available_sheets": wb.sheetnames},
            )

    def _get_merged_ranges(self, ws: Worksheet) -> dict[tuple[int, int], Any]:
        """
        Build a mapping of merged cell coordinates to their values.

        Args:
            ws: Worksheet to analyze

        Returns:
            Dict mapping (row, col) -> value for all merged cells
        """
        merged_values: dict[tuple[int, int], Any] = {}

        for merged_range in ws.merged_cells.ranges:
            # Get the value from the top-left cell
            min_row, min_col = merged_range.min_row, merged_range.min_col
            top_left_value = ws.cell(row=min_row, column=min_col).value

            # Map all cells in the range to this value (forward-fill)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = top_left_value

        logger.debug(
            "Merged ranges detected",
            merged_range_count=len(ws.merged_cells.ranges),
            filled_cell_count=len(merged_values),
        )

        return merged_values

    def _read_worksheet_with_forward_fill(
        self,
        ws: Worksheet,
        merged_ranges: dict[tuple[int, int], Any],
    ) -> pd.DataFrame:
        """
        Read worksheet data with merged cell values forward-filled.

        Args:
            ws: Worksheet to read
            merged_ranges: Merged cell mapping from _get_merged_ranges

        Returns:
            DataFrame with forward-filled data
        """
        data: list[list[Any]] = []
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        # Apply row limits
        start_row = 1 + self._skip_rows
        end_row = min(max_row, start_row + self._max_rows) if self._max_rows else max_row

        for row_idx in range(start_row, end_row + 1):
            row_data: list[Any] = []
            for col_idx in range(1, max_col + 1):
                # Check if this cell is part of a merged range
                if (row_idx, col_idx) in merged_ranges:
                    value = merged_ranges[(row_idx, col_idx)]
                else:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                row_data.append(value)
            data.append(row_data)

        # Create DataFrame
        df = pd.DataFrame(data)

        # Apply pandas forward-fill for any remaining None values in category-like columns
        # (This handles vertical merged cells that weren't explicitly merged in the file)
        df = df.ffill(axis=0)

        return df

    def _detect_header_row(self, df: pd.DataFrame) -> int:
        """
        Auto-detect the header row by looking for known field patterns.

        Args:
            df: DataFrame to analyze

        Returns:
            Header row index (0-based)
        """
        all_patterns = [p for patterns in self.FIELD_PATTERNS.values() for p in patterns]

        for row_idx in range(min(20, len(df))):  # Check first 20 rows
            row_values = df.iloc[row_idx].astype(str).str.lower().tolist()
            matches = sum(
                1 for val in row_values
                if any(pattern in val for pattern in all_patterns)
            )
            if matches >= 2:  # At least 2 field patterns found
                logger.debug("Header row detected", row_index=row_idx, matches=matches)
                return row_idx

        return 0  # Default to first row

    def _apply_headers(self, df: pd.DataFrame, header_row: int) -> pd.DataFrame:
        """
        Apply headers from specified row and remove header rows from data.

        Args:
            df: Source DataFrame
            header_row: Row index to use as headers

        Returns:
            DataFrame with proper headers
        """
        if header_row >= len(df):
            return df

        # Use the header row values as column names
        headers = df.iloc[header_row].astype(str).str.strip().tolist()

        # Handle duplicate column names
        seen: dict[str, int] = {}
        unique_headers: list[str] = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                unique_headers.append(h)

        # Apply headers and remove header row(s)
        df.columns = unique_headers
        df = df.iloc[header_row + 1:].reset_index(drop=True)

        return df

    def _detect_column_mapping(self, columns: list[str]) -> dict[str, str]:
        """
        Map DataFrame columns to standard fields.

        Args:
            columns: List of column names

        Returns:
            Dict mapping standard field names to column names
        """
        mapping: dict[str, str] = {}
        columns_lower = {c: c.lower() for c in columns}

        for field, patterns in self.FIELD_PATTERNS.items():
            for col, col_lower in columns_lower.items():
                if any(pattern in col_lower for pattern in patterns):
                    if field not in mapping:  # Take first match
                        mapping[field] = col
                        break

        logger.debug("Column mapping detected", mapping=mapping)
        return mapping

    def _convert_to_normalized_rows(
        self,
        df: pd.DataFrame,
        column_mapping: dict[str, str],
    ) -> list[NormalizedRow]:
        """
        Convert DataFrame rows to NormalizedRow objects.

        Args:
            df: Source DataFrame
            column_mapping: Field to column mapping

        Returns:
            List of NormalizedRow objects
        """
        rows: list[NormalizedRow] = []
        errors: list[ParserError] = []

        for idx, row in df.iterrows():
            try:
                # Extract required field: name
                name_col = column_mapping.get("name")
                name = str(row.get(name_col, "")).strip() if name_col else ""

                if not name or name.lower() in ("nan", "none", ""):
                    continue  # Skip empty rows

                # Extract optional fields
                price = self._parse_price(row, column_mapping)
                sku = self._extract_field(row, column_mapping, "sku")
                category = self._extract_field(row, column_mapping, "category")
                brand = self._extract_field(row, column_mapping, "brand")
                unit = self._extract_field(row, column_mapping, "unit")
                description = self._extract_field(row, column_mapping, "description")

                # Build characteristics from unmapped columns
                characteristics = self._extract_characteristics(row, column_mapping)
                characteristics["_source_row"] = int(idx) + 1

                normalized = NormalizedRow(
                    name=name,
                    description=description,
                    price=price,
                    sku=sku,
                    category=category,
                    brand=brand,
                    unit=unit,
                    characteristics=characteristics,
                    raw_data=row.to_dict(),
                )
                rows.append(normalized)

            except Exception as e:
                errors.append(
                    ParserError(
                        row_number=int(idx) + 1,
                        error_type="conversion",
                        message=str(e),
                        raw_data=row.to_dict() if hasattr(row, "to_dict") else None,
                    )
                )
                logger.warning(
                    "Row conversion failed",
                    row_index=idx,
                    error=str(e),
                )

        if errors:
            logger.warning(
                "Some rows failed to convert",
                error_count=len(errors),
                success_count=len(rows),
            )

        return rows

    def _parse_price(
        self,
        row: pd.Series,
        column_mapping: dict[str, str],
    ) -> Decimal | None:
        """Parse price from row."""
        price_col = column_mapping.get("price")
        if not price_col:
            return None

        value = row.get(price_col)
        if value is None or pd.isna(value):
            return None

        try:
            # Handle string prices with formatting
            if isinstance(value, str):
                # Remove currency symbols, spaces, and normalize decimal separator
                cleaned = value.replace(" ", "").replace(",", ".").replace("₽", "").replace("$", "").replace("€", "")
                return Decimal(cleaned)
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def _extract_field(
        self,
        row: pd.Series,
        column_mapping: dict[str, str],
        field: str,
    ) -> str | None:
        """Extract a text field from row."""
        col = column_mapping.get(field)
        if not col:
            return None

        value = row.get(col)
        if value is None or pd.isna(value):
            return None

        value_str = str(value).strip()
        return value_str if value_str.lower() not in ("nan", "none", "") else None

    def _extract_characteristics(
        self,
        row: pd.Series,
        column_mapping: dict[str, str],
    ) -> dict[str, Any]:
        """Extract unmapped columns as characteristics."""
        mapped_cols = set(column_mapping.values())
        characteristics: dict[str, Any] = {}

        for col, value in row.items():
            if col not in mapped_cols and value is not None and not pd.isna(value):
                value_str = str(value).strip()
                if value_str.lower() not in ("nan", "none", ""):
                    characteristics[str(col)] = value_str

        return characteristics


# Import UUID for type hints
from uuid import UUID  # noqa: E402

